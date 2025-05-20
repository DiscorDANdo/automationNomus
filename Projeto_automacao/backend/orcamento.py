from openpyxl import load_workbook
from tkinter_class import tkinter_class
from datetime import datetime
from icecream import ic
import pandas as pd
import tabula

class orcamento:
    
    def __init__(self, caminho_pdf, caminho_planilha):
        self.arquivo_pdf = caminho_pdf
        self.arquivo_planilha = caminho_planilha
        self.dados = []
        self.dados_excel = []
        self.peca_excel = {"Código" : "", "Quantidade" : "", "Valor" : ""}

    def extrair_dados(self):
        dfs = tabula.read_pdf(self.arquivo_pdf, pages='all', multiple_tables=True, lattice=True)
        
        dfs_completed = pd.concat(dfs, ignore_index=True)

        df = pd.DataFrame(dfs_completed)
        book = load_workbook(self.arquivo_planilha)
        
        sheet = book["Carbono"]

        linha_inicial = 2

        for i, row in df.iterrows():
            linha_excel = linha_inicial + i
            sheet[f"A{linha_excel}"] = row["Peça"]
            sheet[f"B{linha_excel}"] = row["Espessura"]
            sheet[f"C{linha_excel}"] = f"{self.formatar_valor(row["Largura"])}"
            sheet[f"D{linha_excel}"] = f"{self.formatar_valor(row["Comprimento"])}"
            sheet[f"G{linha_excel}"] = f"{self.formatar_tempo(row["Tempo"])}"
            sheet[f"J{linha_excel}"] = int(row["QNTD"])
            sheet[f"H{linha_excel}"] = row["DOBRA"]
        book.save(self.arquivo_planilha)

    def extrair_dados_excel(self):
            
        input("Abra a planilha, salve e a feche para prosseguir.")
        
        book =  load_workbook(self.arquivo_planilha, data_only=True)
        sheet = book["Carbono"]

        df = pd.read_excel(self.arquivo_planilha, sheet_name="Carbono")

        linha = 2
        
        for i, row in df.iterrows():
            try:
                self.peca_excel["Código"] = row["PEÇA"]
                self.peca_excel["Quantidade"] = row["QTDE"]
                if sheet[f"K{linha}"].value is not None:
                    valor = sheet[f"K{linha}"].value

                valor_formatado = f"{valor:.2f}"
                self.peca_excel["Valor"] = valor_formatado

                if sheet[f"A{linha}"].value is not None:
                    self.dados_excel.append(self.peca_excel.copy())
                    
                linha += 1
            except Exception as e:
                ic(f"Error extracting data: {e}")

        ic("Products successfully extracted from Excel")

    def formatar_valor(self, valor):
        try:
            valor = valor.split(",")[0]
            valor = valor.replace(",", "")

            if len(valor) >= 4:
                primeiro_digito = valor[0]
                restante_digitos = valor[1:]
                valor = f"{primeiro_digito},{restante_digitos}"
                ic(f"Data successfully formatted: {valor}")
            if len(valor) == 3:
                valor = f"0,{valor}"
                ic(f"Data succesfully formatted: {valor}")
            if len(valor) == 2:
                valor = f"0,0{valor}"
                ic(f"Data succesfully formatted: {valor}")
            if len(valor) == 1:
                valor = f"0,00{valor}"
                ic(f"Data succesfully formatted: {valor}")
            
            return valor
        
        except Exception as e:
            ic(f"Error formatting data: {e}")
            return "0,00"
        
    def formatar_tempo(self, tempo):
        try:
            tempo_str = tempo
            tempo = datetime.strptime(tempo_str, "%H:%M:%S")
            tempo_formatado = tempo.hour * 3600 + tempo.minute * 60 + tempo.second
            ic(f"Time successfully formatted: {tempo_formatado}")
            return tempo_formatado
        except Exception as e:
            ic(f"Error formatting time: {e}")
            return 0

    def main():
        ic("Starting data extraction...")
        caminho_pdf = tkinter_class.escolher_pdf()
        caminho_planilha = tkinter_class.escolher_planilha()
        automacao = orcamento(caminho_pdf, caminho_planilha)
        automacao.extrair_dados()
        automacao.extrair_dados_excel()
        ic(automacao.dados_excel[0])
        ic("Data extraction completed successfully.")

if __name__ == "__main__":
    orcamento.main()